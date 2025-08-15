from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError
from typing import List, Optional
from datetime import datetime, timedelta
import json
import os
import tempfile
from calendar import monthrange
from sqlalchemy import func, extract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

from ..core.database import get_db
from ..core.security import get_current_active_user
from ..models.user import User
from ..models.expense import Expense, BusinessUnit, Truck, Trailer, FuelStation, CompanyEnum, ExpenseCategoryEnum
from ..schemas.expense import (
    ExpenseCreate, ExpenseUpdate, Expense as ExpenseSchema,
    BusinessUnitCreate, BusinessUnitUpdate, BusinessUnit as BusinessUnitSchema,
    TruckCreate, TruckUpdate, Truck as TruckSchema,
    TrailerCreate, TrailerUpdate, Trailer as TrailerSchema,
    FuelStationCreate, FuelStationUpdate, FuelStation as FuelStationSchema
)
from ..utils.file_handler import file_handler

router = APIRouter()

def validate_expense_data(expense_data: dict, db: Session) -> dict:
    """Validate expense data including date, price, and description requirements."""
    # Validate date
    if not expense_data.get('date'):
        raise HTTPException(status_code=400, detail="Date is required")
    
    try:
        date_obj = datetime.fromisoformat(expense_data['date'].replace('Z', '+00:00'))
        # Check if date is not in the future
        if date_obj > datetime.now():
            raise HTTPException(status_code=400, detail="Date cannot be in the future")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Validate price
    price = expense_data.get('price')
    if price is None:
        raise HTTPException(status_code=400, detail="Price is required")
    
    try:
        price_float = float(price)
        if price_float <= 0:
            raise HTTPException(status_code=400, detail="Price must be greater than 0")
        expense_data['price'] = price_float  # Ensure price is float
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Price must be a valid number")
    
    # Validate description is provided
    if not expense_data.get('description') or str(expense_data['description']).strip() == '':
        raise HTTPException(status_code=400, detail="Description is required")
    
    return expense_data

def check_duplicate_entity(db: Session, model, field_name: str, value: str, exclude_id: int = None) -> bool:
    """Check if an entity with the same field value already exists."""
    query = db.query(model).filter(getattr(model, field_name) == value)
    if exclude_id:
        query = query.filter(model.id != exclude_id)
    return query.first() is not None

def serialize_expense_with_relationships(expense: Expense) -> dict:
    """
    Serialize expense with relationships to camelCase format for frontend compatibility.
    Centralizes the mapping logic to avoid duplication.
    """
    return {
        "id": expense.id,
        "company": expense.company,
        "category": expense.category,
        "date": expense.date,
        "price": expense.price,
        "description": expense.description,
        "repair_description": expense.repair_description,
        "gallons": expense.gallons,
        "business_unit_id": expense.business_unit_id,
        "truck_id": expense.truck_id,
        "trailer_id": expense.trailer_id,
        "fuel_station_id": expense.fuel_station_id,
        "attachment_path": expense.attachment_path,
        "businessUnit": {
            "id": expense.business_unit.id, 
            "name": expense.business_unit.name
        } if expense.business_unit else None,
        "truck": {
            "id": expense.truck.id, 
            "number": expense.truck.number
        } if expense.truck else None,
        "trailer": {
            "id": expense.trailer.id, 
            "number": expense.trailer.number
        } if expense.trailer else None,
        "fuelStation": {
            "id": expense.fuel_station.id, 
            "name": expense.fuel_station.name
        } if expense.fuel_station else None,
        "created_at": expense.created_at,
        "updated_at": expense.updated_at
    }

def get_expense_with_relationships(db: Session, expense_id: int) -> Expense:
    """
    Get expense with all relationships loaded.
    Centralizes the query logic to avoid duplication.
    """
    return db.query(Expense).options(
        joinedload(Expense.business_unit),
        joinedload(Expense.truck),
        joinedload(Expense.trailer),
        joinedload(Expense.fuel_station)
    ).filter(Expense.id == expense_id).first()

def get_expenses_with_relationships(db: Session, company: Optional[CompanyEnum] = None, 
                                  category: Optional[ExpenseCategoryEnum] = None, 
                                  skip: int = 0, limit: int = 100) -> List[Expense]:
    """
    Get expenses with all relationships loaded and optional filtering.
    Centralizes the query logic to avoid duplication.
    """
    query = db.query(Expense).options(
        joinedload(Expense.business_unit),
        joinedload(Expense.truck),
        joinedload(Expense.trailer),
        joinedload(Expense.fuel_station)
    )
    
    if company:
        query = query.filter(Expense.company == company)
    if category:
        query = query.filter(Expense.category == category)
    
    return query.offset(skip).limit(limit).all()

def check_entity_usage_and_delete(db: Session, entity, entity_id: int, entity_name: str):
    """
    Check if management entity is referenced by expenses and delete if not.
    Centralizes the deletion logic with referential integrity checks.
    """
    expense_count = db.query(Expense).filter(getattr(Expense, f"{entity_name}_id") == entity_id).count()
    if expense_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete {entity_name}: {expense_count} expense(s) reference it"
        )
    
    db.delete(entity)
    db.commit()

# Expense endpoints
@router.post("/expenses/", response_model=dict, status_code=status.HTTP_201_CREATED)
async def create_expense(
    expense_data: str = Form(...),
    attachment: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Create a new expense entry with optional file attachment."""
    try:
        # Parse JSON data from form
        expense_dict = json.loads(expense_data)
        
        # Validate expense data
        expense_dict = validate_expense_data(expense_dict, db)
        
        expense = ExpenseCreate(**expense_dict)
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(status_code=400, detail=f"Invalid expense data: {str(e)}")
    
    # Handle file upload if provided
    attachment_path = None
    if attachment:
        attachment_path = await file_handler.save_file(attachment)
    
    try:
        # Create expense with attachment path
        expense_data_dict = expense.model_dump()
        expense_data_dict["attachment_path"] = attachment_path
        
        db_expense = Expense(**expense_data_dict)
        db.add(db_expense)
        db.commit()
        db.refresh(db_expense)
        
        # Return with relationships
        expense_with_relationships = get_expense_with_relationships(db, db_expense.id)
        response_data = serialize_expense_with_relationships(expense_with_relationships)
        response_data["message"] = "Expense created successfully"
        response_data["status"] = "success"
        return response_data
        
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Database integrity error - check for duplicate values")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create expense: {str(e)}")

@router.get("/expenses/", response_model=List[dict])
async def read_expenses(
    company: Optional[CompanyEnum] = None,
    category: Optional[ExpenseCategoryEnum] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Get all expenses with optional filtering by company and category.
    Returns expenses with relationships serialized for frontend compatibility.
    """
    expenses = get_expenses_with_relationships(db, company, category, skip, limit)
    return [serialize_expense_with_relationships(expense) for expense in expenses]

@router.get("/expenses/{expense_id}", response_model=dict)
def read_expense(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Get a single expense by ID with all relationships."""
    expense = get_expense_with_relationships(db, expense_id)
    
    if expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return serialize_expense_with_relationships(expense)

@router.put("/expenses/{expense_id}", response_model=dict)
async def update_expense(
    expense_id: int,
    expense_data: str = Form(...),
    attachment: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Update an expense by ID with optional file attachment."""
    db_expense = get_expense_with_relationships(db, expense_id)
    if db_expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    try:
        # Parse JSON data from form
        expense_dict = json.loads(expense_data)
        
        # Validate expense data
        expense_dict = validate_expense_data(expense_dict, db)
        
        expense = ExpenseUpdate(**expense_dict)
    except (json.JSONDecodeError, ValueError) as e:
        raise HTTPException(status_code=400, detail=f"Invalid expense data: {str(e)}")
    
    try:
        # Handle file upload if provided
        if attachment:
            # Delete old attachment if it exists
            if db_expense.attachment_path:
                file_handler.delete_file(db_expense.attachment_path)
            
            # Save new attachment
            attachment_path = await file_handler.save_file(attachment)
            expense.attachment_path = attachment_path
        elif attachment is None and 'attachment_path' not in expense_dict:
            # If no file provided and no attachment_path in data, remove existing attachment
            if db_expense.attachment_path:
                file_handler.delete_file(db_expense.attachment_path)
                expense.attachment_path = None
        
        # Update expense
        update_data = expense.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_expense, field, value)
        
        db.commit()
        db.refresh(db_expense)
        
        response_data = serialize_expense_with_relationships(db_expense)
        response_data["message"] = "Expense updated successfully"
        response_data["status"] = "success"
        return response_data
        
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Database integrity error - check for duplicate values")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update expense: {str(e)}")

@router.delete("/expenses/{expense_id}", response_model=dict)
def delete_expense(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Delete an expense by ID with confirmation message."""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    try:
        # Delete associated file if it exists
        if expense.attachment_path:
            file_handler.delete_file(expense.attachment_path)
        
        db.delete(expense)
        db.commit()
        
        return {
            "message": f"Expense with ID {expense_id} deleted successfully", 
            "status": "success",
            "deleted_id": expense_id
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete expense: {str(e)}")

@router.get("/expenses/{expense_id}/attachment")
async def download_attachment(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Download expense attachment file."""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if not expense.attachment_path:
        raise HTTPException(status_code=404, detail="No attachment found for this expense")
    
    file_path = file_handler.get_absolute_path(expense.attachment_path)
    if not file_path or not file_path.exists():
        raise HTTPException(status_code=404, detail="Attachment file not found")
    
    # Get the original filename from the path
    filename = os.path.basename(expense.attachment_path)
    
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type='application/octet-stream'
    )

@router.delete("/expenses/{expense_id}/attachment", status_code=status.HTTP_204_NO_CONTENT)
async def remove_attachment(
    expense_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Remove attachment from expense."""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if expense is None:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if not expense.attachment_path:
        raise HTTPException(status_code=404, detail="No attachment found for this expense")
    
    # Delete the file
    file_handler.delete_file(expense.attachment_path)
    
    # Remove path from database
    expense.attachment_path = None
    db.commit()
    
    return

# Business Unit endpoints
@router.post("/business-units/", response_model=dict, status_code=status.HTTP_201_CREATED)
def create_business_unit(
    business_unit: BusinessUnitCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Create a new business unit with duplicate checking."""
    # Check for duplicate name
    if check_duplicate_entity(db, BusinessUnit, "name", business_unit.name):
        raise HTTPException(status_code=400, detail=f"Business unit with name '{business_unit.name}' already exists")
    
    try:
        db_business_unit = BusinessUnit(**business_unit.model_dump())
        db.add(db_business_unit)
        db.commit()
        db.refresh(db_business_unit)
        
        return {
            **business_unit.model_dump(),
            "id": db_business_unit.id,
            "created_at": db_business_unit.created_at,
            "updated_at": db_business_unit.updated_at,
            "message": "Business unit created successfully",
            "status": "success"
        }
        
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Business unit name must be unique")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create business unit: {str(e)}")

@router.get("/business-units/", response_model=List[BusinessUnitSchema])
async def read_business_units(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    business_units = db.query(BusinessUnit).offset(skip).limit(limit).all()
    return business_units

@router.put("/business-units/{business_unit_id}", response_model=BusinessUnitSchema)
def update_business_unit(
    business_unit_id: int,
    business_unit: BusinessUnitCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    db_business_unit = db.query(BusinessUnit).filter(BusinessUnit.id == business_unit_id).first()
    if not db_business_unit:
        raise HTTPException(status_code=404, detail="Business unit not found")
    
    for key, value in business_unit.model_dump().items():
        setattr(db_business_unit, key, value)
    
    db.commit()
    db.refresh(db_business_unit)
    return db_business_unit

@router.delete("/business-units/{business_unit_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_business_unit(
    business_unit_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Delete a business unit if not referenced by expenses."""
    business_unit = db.query(BusinessUnit).filter(BusinessUnit.id == business_unit_id).first()
    if not business_unit:
        raise HTTPException(status_code=404, detail="Business unit not found")
    
    check_entity_usage_and_delete(db, business_unit, business_unit_id, "business_unit")

# Truck endpoints
@router.post("/trucks/", response_model=dict, status_code=status.HTTP_201_CREATED)
def create_truck(
    truck: TruckCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Create a new truck with duplicate number checking."""
    # Check for duplicate number
    if check_duplicate_entity(db, Truck, "number", truck.number):
        raise HTTPException(status_code=400, detail=f"Truck with number '{truck.number}' already exists")
    
    try:
        db_truck = Truck(**truck.model_dump())
        db.add(db_truck)
        db.commit()
        db.refresh(db_truck)
        
        return {
            **truck.model_dump(),
            "id": db_truck.id,
            "created_at": db_truck.created_at,
            "updated_at": db_truck.updated_at,
            "message": "Truck created successfully",
            "status": "success"
        }
        
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Truck number must be unique")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create truck: {str(e)}")

@router.get("/trucks/", response_model=List[TruckSchema])
async def read_trucks(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    trucks = db.query(Truck).offset(skip).limit(limit).all()
    return trucks

@router.put("/trucks/{truck_id}", response_model=TruckSchema)
def update_truck(
    truck_id: int,
    truck: TruckCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    db_truck = db.query(Truck).filter(Truck.id == truck_id).first()
    if not db_truck:
        raise HTTPException(status_code=404, detail="Truck not found")
    
    for key, value in truck.model_dump().items():
        setattr(db_truck, key, value)
    
    db.commit()
    db.refresh(db_truck)
    return db_truck

@router.delete("/trucks/{truck_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_truck(
    truck_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Delete a truck if not referenced by expenses."""
    truck = db.query(Truck).filter(Truck.id == truck_id).first()
    if not truck:
        raise HTTPException(status_code=404, detail="Truck not found")
    
    check_entity_usage_and_delete(db, truck, truck_id, "truck")

# Trailer endpoints
@router.post("/trailers/", response_model=dict, status_code=status.HTTP_201_CREATED)
def create_trailer(
    trailer: TrailerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Create a new trailer with duplicate number checking."""
    # Check for duplicate number
    if check_duplicate_entity(db, Trailer, "number", trailer.number):
        raise HTTPException(status_code=400, detail=f"Trailer with number '{trailer.number}' already exists")
    
    try:
        db_trailer = Trailer(**trailer.model_dump())
        db.add(db_trailer)
        db.commit()
        db.refresh(db_trailer)
        
        return {
            **trailer.model_dump(),
            "id": db_trailer.id,
            "created_at": db_trailer.created_at,
            "updated_at": db_trailer.updated_at,
            "message": "Trailer created successfully",
            "status": "success"
        }
        
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Trailer number must be unique")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create trailer: {str(e)}")

@router.get("/trailers/", response_model=List[TrailerSchema])
async def read_trailers(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    trailers = db.query(Trailer).offset(skip).limit(limit).all()
    return trailers

@router.put("/trailers/{trailer_id}", response_model=TrailerSchema)
def update_trailer(
    trailer_id: int,
    trailer: TrailerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    db_trailer = db.query(Trailer).filter(Trailer.id == trailer_id).first()
    if not db_trailer:
        raise HTTPException(status_code=404, detail="Trailer not found")
    
    for key, value in trailer.model_dump().items():
        setattr(db_trailer, key, value)
    
    db.commit()
    db.refresh(db_trailer)
    return db_trailer

@router.delete("/trailers/{trailer_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_trailer(
    trailer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Delete a trailer if not referenced by expenses."""
    trailer = db.query(Trailer).filter(Trailer.id == trailer_id).first()
    if not trailer:
        raise HTTPException(status_code=404, detail="Trailer not found")
    
    check_entity_usage_and_delete(db, trailer, trailer_id, "trailer")

# Fuel Station endpoints
@router.post("/fuel-stations/", response_model=dict, status_code=status.HTTP_201_CREATED)
def create_fuel_station(
    fuel_station: FuelStationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Create a new fuel station with duplicate name checking."""
    # Check for duplicate name
    if check_duplicate_entity(db, FuelStation, "name", fuel_station.name):
        raise HTTPException(status_code=400, detail=f"Fuel station with name '{fuel_station.name}' already exists")
    
    try:
        db_fuel_station = FuelStation(**fuel_station.model_dump())
        db.add(db_fuel_station)
        db.commit()
        db.refresh(db_fuel_station)
        
        return {
            **fuel_station.model_dump(),
            "id": db_fuel_station.id,
            "created_at": db_fuel_station.created_at,
            "updated_at": db_fuel_station.updated_at,
            "message": "Fuel station created successfully",
            "status": "success"
        }
        
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Fuel station name must be unique")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create fuel station: {str(e)}")

@router.get("/fuel-stations/", response_model=List[FuelStationSchema])
async def read_fuel_stations(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    fuel_stations = db.query(FuelStation).offset(skip).limit(limit).all()
    return fuel_stations

@router.put("/fuel-stations/{fuel_station_id}", response_model=FuelStationSchema)
def update_fuel_station(
    fuel_station_id: int,
    fuel_station: FuelStationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    db_fuel_station = db.query(FuelStation).filter(FuelStation.id == fuel_station_id).first()
    if not db_fuel_station:
        raise HTTPException(status_code=404, detail="Fuel station not found")
    
    for key, value in fuel_station.model_dump().items():
        setattr(db_fuel_station, key, value)
    
    db.commit()
    db.refresh(db_fuel_station)
    return db_fuel_station

@router.delete("/fuel-stations/{fuel_station_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_fuel_station(
    fuel_station_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Delete a fuel station if not referenced by expenses."""
    fuel_station = db.query(FuelStation).filter(FuelStation.id == fuel_station_id).first()
    if not fuel_station:
        raise HTTPException(status_code=404, detail="Fuel station not found")
    
    check_entity_usage_and_delete(db, fuel_station, fuel_station_id, "fuel_station")

# Analytics endpoints
@router.get("/analytics/monthly-change/{company}")
def get_monthly_change(
    company: CompanyEnum,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Get month-over-month percentage change in expenses for a company."""
    try:
        now = datetime.now()
        
        # Current month
        current_month_start = datetime(now.year, now.month, 1)
        current_month_end = datetime(now.year, now.month, monthrange(now.year, now.month)[1], 23, 59, 59)
        
        # Previous month
        if now.month == 1:
            prev_month = 12
            prev_year = now.year - 1
        else:
            prev_month = now.month - 1
            prev_year = now.year
        
        prev_month_start = datetime(prev_year, prev_month, 1)
        prev_month_end = datetime(prev_year, prev_month, monthrange(prev_year, prev_month)[1], 23, 59, 59)
        
        # Calculate current month total
        current_total = db.query(func.sum(Expense.price)).filter(
            Expense.company == company,
            Expense.date >= current_month_start,
            Expense.date <= current_month_end
        ).scalar() or 0
        
        # Calculate previous month total
        prev_total = db.query(func.sum(Expense.price)).filter(
            Expense.company == company,
            Expense.date >= prev_month_start,
            Expense.date <= prev_month_end
        ).scalar() or 0
        
        # Calculate percentage change
        if prev_total > 0:
            percentage_change = ((current_total - prev_total) / prev_total) * 100
        else:
            percentage_change = 0 if current_total == 0 else 100
        
        # Get monthly data for the last 6 months for trend
        monthly_data = []
        for i in range(5, -1, -1):  # Last 6 months including current
            if now.month - i <= 0:
                month = 12 + (now.month - i)
                year = now.year - 1
            else:
                month = now.month - i
                year = now.year
            
            month_start = datetime(year, month, 1)
            month_end = datetime(year, month, monthrange(year, month)[1], 23, 59, 59)
            
            month_total = db.query(func.sum(Expense.price)).filter(
                Expense.company == company,
                Expense.date >= month_start,
                Expense.date <= month_end
            ).scalar() or 0
            
            monthly_data.append({
                "month": datetime(year, month, 1).strftime('%b %Y'),
                "total": float(month_total)
            })
        
        return {
            "current_month": float(current_total),
            "previous_month": float(prev_total),
            "percentage_change": round(percentage_change, 2),
            "monthly_trend": monthly_data
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analytics failed: {str(e)}")

@router.get("/analytics/top-categories/{company}")
def get_top_categories(
    company: CompanyEnum,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Get top 3 expense categories with monthly trends for a company."""
    try:
        # Get total by category for the last 6 months
        now = datetime.now()
        six_months_ago = now - timedelta(days=180)
        
        # Get category totals
        category_totals = db.query(
            Expense.category,
            func.sum(Expense.price).label('total')
        ).filter(
            Expense.company == company,
            Expense.date >= six_months_ago
        ).group_by(Expense.category).order_by(func.sum(Expense.price).desc()).limit(3).all()
        
        # Get monthly trends for each top category
        top_categories_data = []
        for category_total in category_totals:
            category = category_total.category
            
            # Get monthly data for this category
            monthly_data = []
            for i in range(5, -1, -1):  # Last 6 months including current
                if now.month - i <= 0:
                    month = 12 + (now.month - i)
                    year = now.year - 1
                else:
                    month = now.month - i
                    year = now.year
                
                month_start = datetime(year, month, 1)
                month_end = datetime(year, month, monthrange(year, month)[1], 23, 59, 59)
                
                month_total = db.query(func.sum(Expense.price)).filter(
                    Expense.company == company,
                    Expense.category == category,
                    Expense.date >= month_start,
                    Expense.date <= month_end
                ).scalar() or 0
                
                monthly_data.append({
                    "month": datetime(year, month, 1).strftime('%b %Y'),
                    "amount": float(month_total)
                })
            
            # Format category name for display
            category_display = category.replace('-', ' ').title()
            
            top_categories_data.append({
                "category": category,
                "category_display": category_display,
                "total": float(category_total.total),
                "monthly_data": monthly_data
            })
        
        return {
            "top_categories": top_categories_data
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analytics failed: {str(e)}")

@router.get("/export/{company}")
def export_company_data(
    company: CompanyEnum,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Export all expense data for a company to Excel format."""
    try:
        # Get all expenses for the company with relationships
        expenses = db.query(Expense).options(
            joinedload(Expense.business_unit),
            joinedload(Expense.truck),
            joinedload(Expense.trailer),
            joinedload(Expense.fuel_station)
        ).filter(Expense.company == company).order_by(Expense.date.desc()).all()

        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define expense categories and their fields
        categories = {
            'truck': ['Date', 'Business Unit', 'Truck Number', 'Repair Description', 'Price ($)'],
            'trailer': ['Date', 'Business Unit', 'Trailer Number', 'Repair Description', 'Price ($)'],
            'dmv': ['Date', 'Description', 'Price ($)'],
            'parts': ['Date', 'Description', 'Price ($)'],
            'phone-tracker': ['Date', 'Description', 'Price ($)'],
            'other-expenses': ['Date', 'Description', 'Price ($)'],
            'toll': ['Date', 'Price ($)'],
            'office-supplies': ['Date', 'Description', 'Price ($)'],
            'fuel-diesel': ['Date', 'Fuel Station', 'Gallons', 'Price ($)'],
            'def': ['Date', 'Price ($)']
        }
        
        # Create a sheet for each category that has data
        for category, fields in categories.items():
            category_expenses = [e for e in expenses if e.category == category]
            if not category_expenses:
                continue
                
            # Create worksheet
            ws = wb.create_sheet(title=category.replace('-', ' ').title())
            
            # Style for headers
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Add headers
            for col, field in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col, value=field)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data rows
            for row, expense in enumerate(category_expenses, 2):
                col = 1
                
                # Date
                ws.cell(row=row, column=col, value=expense.date.strftime('%Y-%m-%d') if expense.date else '')
                col += 1
                
                # Category specific fields
                if category in ['truck', 'trailer']:
                    # Business Unit
                    ws.cell(row=row, column=col, value=expense.business_unit.name if expense.business_unit else '')
                    col += 1
                    
                    # Truck/Trailer Number
                    if category == 'truck':
                        ws.cell(row=row, column=col, value=expense.truck.number if expense.truck else '')
                    else:
                        ws.cell(row=row, column=col, value=expense.trailer.number if expense.trailer else '')
                    col += 1
                    
                    # Repair Description
                    ws.cell(row=row, column=col, value=expense.repair_description or '')
                    col += 1
                    
                elif category == 'fuel-diesel':
                    # Fuel Station
                    ws.cell(row=row, column=col, value=expense.fuel_station.name if expense.fuel_station else '')
                    col += 1
                    
                    # Gallons
                    ws.cell(row=row, column=col, value=expense.gallons or '')
                    col += 1
                    
                elif category not in ['toll', 'def']:
                    # Description for categories that have it
                    ws.cell(row=row, column=col, value=expense.description or '')
                    col += 1
                
                # Price (always last column) - format as currency
                price_cell = ws.cell(row=row, column=col, value=float(expense.price) if expense.price else 0)
                price_cell.number_format = '"$"#,##0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        if expenses:
            summary_ws = wb.create_sheet(title='Summary', index=0)
            
            # Headers
            summary_ws.cell(row=1, column=1, value='Category').font = header_font
            summary_ws.cell(row=1, column=1).fill = header_fill
            summary_ws.cell(row=1, column=2, value='Total Expenses').font = header_font
            summary_ws.cell(row=1, column=2).fill = header_fill
            summary_ws.cell(row=1, column=3, value='Total Price ($)').font = header_font
            summary_ws.cell(row=1, column=3).fill = header_fill
            
            # Calculate summary data
            category_totals = {}
            for expense in expenses:
                category = expense.category
                if category not in category_totals:
                    category_totals[category] = {'count': 0, 'total': 0}
                category_totals[category]['count'] += 1
                category_totals[category]['total'] += float(expense.price) if expense.price else 0
            
            # Add summary data
            row = 2
            grand_total = 0
            total_expenses = 0
            for category, data in category_totals.items():
                summary_ws.cell(row=row, column=1, value=category.replace('-', ' ').title())
                summary_ws.cell(row=row, column=2, value=data['count'])
                total_cell = summary_ws.cell(row=row, column=3, value=data['total'])
                total_cell.number_format = '"$"#,##0.00'
                grand_total += data['total']
                total_expenses += data['count']
                row += 1
            
            # Add grand total
            row += 1
            summary_ws.cell(row=row, column=1, value='GRAND TOTAL').font = Font(bold=True)
            summary_ws.cell(row=row, column=2, value=total_expenses).font = Font(bold=True)
            grand_total_cell = summary_ws.cell(row=row, column=3, value=grand_total)
            grand_total_cell.font = Font(bold=True)
            grand_total_cell.number_format = '"$"#,##0.00'
            
            # Auto-adjust column widths for summary
            for column in summary_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                summary_ws.column_dimensions[column_letter].width = adjusted_width

        # If no expenses, create an empty summary sheet
        if not wb.worksheets:
            ws = wb.create_sheet(title='No Data')
            ws.cell(row=1, column=1, value='No expense data found for this company')
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        # Generate filename
        company_name = company.value.upper()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{company_name}_Expenses_Export_{timestamp}.xlsx"
        
        # Return file response
        return FileResponse(
            path=tmp_file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.get("/pie-chart-data/{company}")
async def get_pie_chart_data(
    company: CompanyEnum,
    period: str = Query("total", pattern="^(this-month|total)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Get pie chart data for expense categories.
    Period can be 'this-month' or 'total'.
    """
    try:
        query = db.query(Expense).filter(Expense.company == company)
        
        # Filter by period if specified
        if period == "this-month":
            now = datetime.now()
            start_of_month = datetime(now.year, now.month, 1)
            query = query.filter(Expense.date >= start_of_month.date())
        
        expenses = query.all()
        
        # Group by category and sum costs
        category_totals = {}
        for expense in expenses:
            category = expense.category.value
            price = float(expense.price or 0)
            
            if category in category_totals:
                category_totals[category] += price
            else:
                category_totals[category] = price
        
        # Convert to pie chart format with default Mantine colors
        category_colors = {
            'fuel_diesel': '#1c7ed6',  # blue
            'vehicle_repair': '#37b24d',  # green
            'inventory': '#f59f00',  # orange
            'payroll': '#e64980',  # pink
            'insurance': '#7950f2',  # violet
            'office_supplies': '#15aabf',  # cyan
            'marketing': '#fd7e14',  # orange.6
            'travel': '#51cf66',  # lime
            'maintenance': '#ffd43b',  # yellow
            'utilities': '#f06292'  # pink.5
        }
        
        pie_data = []
        for category, total in category_totals.items():
            if total > 0:  # Only include categories with expenses
                # Convert category key to display name
                display_name = category.replace('_', ' ').title()
                pie_data.append({
                    "category": category,
                    "name": display_name,
                    "value": round(total, 2),
                    "color": category_colors.get(category, '#868e96')  # default gray
                })
        
        # Sort by value descending
        pie_data.sort(key=lambda x: x['value'], reverse=True)
        
        return {
            "company": company.value,
            "period": period,
            "data": pie_data,
            "total_amount": round(sum(item['value'] for item in pie_data), 2),
            "category_count": len(pie_data)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get pie chart data: {str(e)}")
